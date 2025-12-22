"""
Rapor Oluşturma Modülü
Excel, PDF ve HTML formatlarında detaylı raporlar üretir
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import seaborn as sns
import sys
import os

# Path ayarlaması
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from src.utils.data_loader import DataLoader
from src.utils.statistics import StatisticalCalculator
from src.analysis.descriptive_stats import DescriptiveAnalysis
from src.analysis.control_charts import ControlCharts
from src.analysis.capability_analysis import ProcessCapability
from src.analysis.pareto_analysis import ParetoAnalysis
from src.visualization.charts import ChartGenerator
# from src.visualization.dashboard import Dashboard  # KALDIRILDI - circular import

class ReportGenerator:
    """
    Profesyonel rapor üretimi için ana sınıf
    """
    
    def __init__(self, data_path='data/raw/DATA_SET_MOTOR.xlsx'):
        """
        Args:
            data_path: Veri dosyası yolu
        """
        self.loader = DataLoader(data_path)
        self.df = self.loader.load_data()
        self.stat_calc = StatisticalCalculator()
        self.descriptive = DescriptiveAnalysis(data_path)
        self.control_charts = ControlCharts(data_path)
        self.capability = ProcessCapability(data_path)
        self.pareto = ParetoAnalysis(data_path)
        self.charts = ChartGenerator(data_path)
        # self.dashboard = Dashboard(data_path)  # KALDIRILDI - circular import
        
        # Rapor klasörünü oluştur
        os.makedirs('reports', exist_ok=True)
        os.makedirs('reports/images', exist_ok=True)
    
    def generate_executive_summary(self, output_path='reports/executive_summary.xlsx'):
        """
        Yönetici özeti raporu (1 sayfalık)
        
        Args:
            output_path: Çıktı dosyası yolu
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Yönetici Özeti"
        
        # Başlık
        ws['A1'] = "MOTOR ÜRETİM KALİTE RAPORU - YÖNETİCİ ÖZETİ"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:F1')
        
        # Tarih
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:F2')
        
        # KPI Başlıkları
        ws['A4'] = "KRİTİK PERFORMANS GÖSTERGELERİ"
        ws['A4'].font = Font(size=12, bold=True, color="2563EB")
        ws.merge_cells('A4:F4')
        
        # KPI Tablosu
        kpi_headers = ['Metrik', 'Mevcut', 'Hedef', 'Durum', 'Fark', 'Trend']
        for col, header in enumerate(kpi_headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # KPI Verileri
        stats = self.loader.get_summary_stats()
        kpi_data = [
            ['OEE (%)', stats['oee'], 85, '❌' if stats['oee'] < 85 else '✅', 
             stats['oee'] - 85, '↓' if stats['oee'] < 80 else '→'],
            ['Verimlilik (%)', round(stats['ortalama_verimlilik'], 1), 90, 
             '❌' if stats['ortalama_verimlilik'] < 90 else '✅', 
             round(stats['ortalama_verimlilik'] - 90, 1), '→'],
            ['Kalite (%)', round(stats['kalite_orani'], 1), 99, 
             '❌' if stats['kalite_orani'] < 99 else '✅', 
             round(stats['kalite_orani'] - 99, 1), '↓'],
            ['Hata Oranı (%)', round(stats['hata_orani'], 1), 2, 
             '❌' if stats['hata_orani'] > 2 else '✅', 
             round(stats['hata_orani'] - 2, 1), '↑']
        ]
        
        for row_idx, kpi in enumerate(kpi_data, 6):
            for col_idx, value in enumerate(kpi, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 4:  # Durum sütunu
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [2, 3, 5]:  # Sayısal sütunlar
                    cell.alignment = Alignment(horizontal="right")
        
        # Ana Bulgular
        ws['A11'] = "ANA BULGULAR"
        ws['A11'].font = Font(size=12, bold=True, color="2563EB")
        ws.merge_cells('A11:F11')
        
        # Bulgular listesi
        findings = [
            f"📊 Toplam {stats['toplam_motor']} motor üretilmiş ({stats['tarih_araligi']})",
            f"⚠️ Hata oranı %{stats['hata_orani']} (Hedef: <%2) - {stats['hatali_motor']} hatalı motor",
            f"📉 OEE %{stats['oee']} seviyesinde (Hedef: %85) - İyileştirme gerekli",
            f"🔧 En sık hata: Sızdırmazlık Hatası (%50 oranında)",
            f"⏱️ Toplam kayıp süre: {stats['toplam_durma'] + stats['toplam_kk_hazirlik'] + stats['toplam_kk_surec']:.0f} saat"
        ]
        
        for idx, finding in enumerate(findings, 12):
            ws[f'A{idx}'] = finding
            ws.merge_cells(f'A{idx}:F{idx}')
        
        # Öneriler
        ws['A18'] = "KRİTİK ÖNERİLER"
        ws['A18'].font = Font(size=12, bold=True, color="EF4444")
        ws.merge_cells('A18:F18')
        
        recommendations = [
            "1. [ACİL] Sızdırmazlık testi prosedürünü gözden geçirin",
            "2. [YÜKSEK] Montaj operatörlerine ek eğitim verin",
            "3. [ORTA] Kalite kontrol süreçlerini otomatikleştirin"
        ]
        
        for idx, rec in enumerate(recommendations, 19):
            ws[f'A{idx}'] = rec
            ws.merge_cells(f'A{idx}:F{idx}')
        
        # Süreç Yeterlilik
        cap = self.capability.calculate_capability_indices('Toplam_Uretim_Suresi')
        
        ws['A23'] = "SÜREÇ YETERLİLİK"
        ws['A23'].font = Font(size=12, bold=True, color="2563EB")
        ws.merge_cells('A23:F23')
        
        ws['A24'] = f"Cpk: {cap['indices']['Cpk']} (Hedef: ≥1.33)"
        ws['B24'] = f"Sigma Seviyesi: {cap['sigma_level']}σ"
        ws['C24'] = f"PPM: {cap['ppm']['expected_total']:,.0f}"
        
        # Format ayarları
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Kenarlık ekle
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=5, max_row=9, min_col=1, max_col=6):
            for cell in row:
                cell.border = border
        
        # Koşullu formatlama - DÜZELTME: Yeni syntax kullanıldı
        if hasattr(ws, 'conditional_formatting'):
            # Excel koşullu formatlama yerine doğrudan renklendirme yapalım
            for row in ws.iter_rows(min_row=6, max_row=9, min_col=5, max_col=5):
                for cell in row:
                    if cell.value and cell.value > 0:
                        cell.font = Font(color='FF0000')  # Kırmızı - Yeni syntax
                    else:
                        cell.font = Font(color='00B050')  # Yeşil - Yeni syntax
        
        # Kaydet
        wb.save(output_path)
        print(f"✅ Yönetici özeti raporu oluşturuldu: {output_path}")
        return output_path
    
    def generate_weekly_report(self, week_number=None, output_path='reports/weekly_report.xlsx'):
        """
        Haftalık detaylı rapor
        
        Args:
            week_number: Hafta numarası (None ise son hafta)
            output_path: Çıktı dosyası
        """
        wb = Workbook()
        
        # 1. Özet Sayfası
        ws_summary = wb.active
        ws_summary.title = "Haftalık Özet"
        
        # Başlık
        ws_summary['A1'] = f"HAFTALIK KALİTE RAPORU - Hafta {week_number or 'Son'}"
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary.merge_cells('A1:H1')
        
        # Hafta verisini filtrele
        if week_number:
            week_data = self.df[self.df['Hafta'] == week_number]
        else:
            week_data = self.df[self.df['Hafta'] == self.df['Hafta'].max()]
        
        # Özet metrikler
        summary_metrics = {
            'Toplam Üretim': len(week_data),
            'Hatalı Motor': week_data['Hatali'].sum(),
            'Hata Oranı (%)': round(week_data['Hatali'].mean() * 100, 2),
            'Ort. Verimlilik (%)': round(week_data['Verimlilik'].mean(), 2),
            'Ort. Üretim Süresi': round(week_data['Toplam_Uretim_Suresi'].mean(), 2),
            'Toplam Durma (saat)': round(week_data['Durma_Suresi_Saat'].sum(), 1)
        }
        
        row = 3
        for metric, value in summary_metrics.items():
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = value
            row += 1
        
        # 2. Günlük Detay Sayfası
        ws_daily = wb.create_sheet("Günlük Detay")
        
        # Günlük aggregasyon
        daily_summary = week_data.groupby('Tarih').agg({
            'Motor_ID': 'count',
            'Verimlilik': 'mean',
            'Hatali': 'sum',
            'Toplam_Uretim_Suresi': 'mean',
            'Durma_Suresi_Saat': 'sum'
        }).round(2)
        
        # DataFrame'i Excel'e yaz
        for r in dataframe_to_rows(daily_summary.reset_index(), index=False, header=True):
            ws_daily.append(r)
        
        # 3. Vardiya Performansı Sayfası
        ws_shift = wb.create_sheet("Vardiya Performansı")
        
        shift_summary = week_data.groupby('Vardiya').agg({
            'Motor_ID': 'count',
            'Verimlilik': ['mean', 'std'],
            'Hatali': 'sum'
        }).round(2)
        
        for r in dataframe_to_rows(shift_summary.reset_index(), index=False, header=True):
            ws_shift.append(r)
        
        # 4. Hata Analizi Sayfası
        ws_defects = wb.create_sheet("Hata Analizi")
        
        defect_summary = week_data[week_data['Hatali'] == 1].groupby('Hata_Nedeni').size().sort_values(ascending=False)
        
        ws_defects['A1'] = "Hata Tipi"
        ws_defects['B1'] = "Adet"
        ws_defects['C1'] = "Yüzde"
        
        for idx, (defect, count) in enumerate(defect_summary.items(), 2):
            ws_defects[f'A{idx}'] = defect
            ws_defects[f'B{idx}'] = count
            ws_defects[f'C{idx}'] = round((count / defect_summary.sum()) * 100, 1)
        
        # Kaydet
        wb.save(output_path)
        print(f"✅ Haftalık rapor oluşturuldu: {output_path}")
        return output_path
    
    def generate_html_report(self, output_path='reports/dashboard_report.html'):
        """
        HTML formatında interaktif rapor
        
        Args:
            output_path: Çıktı HTML dosyası
        """
        # Dashboard verilerini al - DÜZELTME: doğrudan charts kullan
        dashboard_data = {
            'kpis': self.charts.get_kpi_cards(),
            'pareto': self.charts.get_pareto_chart_data(),
            'capability': self.capability.calculate_capability_indices('Toplam_Uretim_Suresi')
        }
        
        html_content = """
        <!DOCTYPE html>
        <html lang="tr">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Motor Üretim Kalite Raporu</title>
            <style>
                * {
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }
                
                body {
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    padding: 20px;
                }
                
                .container {
                    max-width: 1400px;
                    margin: 0 auto;
                    background: rgba(255, 255, 255, 0.95);
                    border-radius: 20px;
                    padding: 30px;
                    box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
                }
                
                h1 {
                    color: #2563eb;
                    margin-bottom: 30px;
                    text-align: center;
                    font-size: 2.5em;
                }
                
                .timestamp {
                    text-align: center;
                    color: #6b7280;
                    margin-bottom: 30px;
                    font-style: italic;
                }
                
                .kpi-grid {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                    gap: 20px;
                    margin-bottom: 40px;
                }
                
                .kpi-card {
                    background: white;
                    border-radius: 12px;
                    padding: 20px;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                    position: relative;
                    overflow: hidden;
                }
                
                .kpi-card::before {
                    content: '';
                    position: absolute;
                    left: 0;
                    top: 0;
                    width: 4px;
                    height: 100%;
                    background: #2563eb;
                }
                
                .kpi-title {
                    color: #6b7280;
                    font-size: 0.9em;
                    text-transform: uppercase;
                    letter-spacing: 0.5px;
                    margin-bottom: 10px;
                }
                
                .kpi-value {
                    font-size: 2.5em;
                    font-weight: bold;
                    color: #1f2937;
                    margin-bottom: 10px;
                }
                
                .kpi-status {
                    font-size: 0.9em;
                    color: #6b7280;
                }
                
                .kpi-status.status-success { color: #10b981; }
                .kpi-status.status-warning { color: #f59e0b; }
                .kpi-status.status-danger { color: #ef4444; }
                
                .section {
                    background: white;
                    border-radius: 12px;
                    padding: 25px;
                    margin-bottom: 30px;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                }
                
                .section-title {
                    color: #2563eb;
                    font-size: 1.5em;
                    margin-bottom: 20px;
                    padding-bottom: 10px;
                    border-bottom: 2px solid #e5e7eb;
                }
                
                table {
                    width: 100%;
                    border-collapse: collapse;
                }
                
                th {
                    background: #f3f4f6;
                    padding: 12px;
                    text-align: left;
                    font-weight: 600;
                    color: #374151;
                }
                
                td {
                    padding: 12px;
                    border-top: 1px solid #e5e7eb;
                }
                
                tr:hover {
                    background: #f9fafb;
                }
                
                .alert {
                    padding: 15px;
                    border-radius: 8px;
                    margin-bottom: 15px;
                    display: flex;
                    align-items: center;
                    gap: 10px;
                }
                
                .alert-danger {
                    background: #fee2e2;
                    border-left: 4px solid #ef4444;
                    color: #991b1b;
                }
                
                .alert-warning {
                    background: #fef3c7;
                    border-left: 4px solid #f59e0b;
                    color: #92400e;
                }
                
                .chart-container {
                    margin: 20px 0;
                    padding: 20px;
                    background: #f9fafb;
                    border-radius: 8px;
                }
                
                @media print {
                    body {
                        background: white;
                    }
                    
                    .container {
                        box-shadow: none;
                    }
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>🏭 Motor Üretim Kalite Raporu</h1>
                <div class="timestamp">
        """
        
        html_content += f"Rapor Tarihi: {datetime.now().strftime('%d %B %Y %H:%M')}</div>"
        
        # KPI Kartları - DÜZELTME: Dictionary'yi doğru iterate et
        html_content += '<div class="kpi-grid">'
        
        for key, kpi in dashboard_data['kpis'].items():
            # Status'u color'dan al
            status = kpi.get('color', 'info')
            html_content += f"""
                    <div class="kpi-card">
                        <div class="kpi-title">{kpi['title']}</div>
                        <div class="kpi-value">{kpi['value']}{kpi['unit']}</div>
                        <div class="kpi-status status-{status}">
                            Hedef: {kpi['target']}{kpi['unit']} | Trend: {kpi.get('trend', 0):+.1f}
                        </div>
                    </div>
            """
        
        html_content += """
                </div>
                
                <!-- Ana Bulgular -->
                <div class="section">
                    <h2 class="section-title">📊 Ana Bulgular</h2>
        """
        
        # Özet istatistikler
        stats = self.loader.get_summary_stats()
        
        html_content += f"""
                    <ul style="line-height: 1.8;">
                        <li>Toplam <strong>{stats['toplam_motor']}</strong> motor üretilmiş</li>
                        <li>Hata oranı: <strong>%{stats['hata_orani']}</strong> ({stats['hatali_motor']} hatalı)</li>
                        <li>Ortalama verimlilik: <strong>%{stats['ortalama_verimlilik']}</strong></li>
                        <li>OEE performansı: <strong>%{stats['oee']}</strong></li>
                    </ul>
                </div>
                
                <!-- Hata Analizi -->
                <div class="section">
                    <h2 class="section-title">🔍 Hata Analizi</h2>
                    <table>
                        <thead>
                            <tr>
                                <th>Hata Tipi</th>
                                <th>Adet</th>
                                <th>Yüzde</th>
                                <th>Kategori</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        # Pareto analizi
        pareto_data = self.loader.get_pareto_data()
        for _, row in pareto_data.iterrows():
            html_content += f"""
                            <tr>
                                <td>{row['Hata_Nedeni']}</td>
                                <td>{row['Adet']}</td>
                                <td>{row['Yuzde']:.1f}%</td>
                                <td>{row['Kategori']}</td>
                            </tr>
            """
        
        html_content += """
                        </tbody>
                    </table>
                </div>
                
                <!-- Süreç Yeterlilik -->
                <div class="section">
                    <h2 class="section-title">📈 Süreç Yeterlilik</h2>
        """
        
        cap = dashboard_data['capability']
        cpk_status = 'danger' if cap['indices']['Cpk'] < 1.0 else 'warning' if cap['indices']['Cpk'] < 1.33 else 'success'
        
        html_content += f"""
                    <div class="kpi-grid">
                        <div class="kpi-card">
                            <div class="kpi-title">Cpk</div>
                            <div class="kpi-value">{cap['indices']['Cpk']}</div>
                            <div class="kpi-status status-{cpk_status}">
                                Hedef: ≥1.33
                            </div>
                        </div>
                        <div class="kpi-card">
                            <div class="kpi-title">Sigma Seviyesi</div>
                            <div class="kpi-value">{cap['sigma_level']}σ</div>
                        </div>
                        <div class="kpi-card">
                            <div class="kpi-title">PPM</div>
                            <div class="kpi-value">{cap['ppm']['expected_total']:,.0f}</div>
                        </div>
                    </div>
                </div>
                
                <!-- Öneriler -->
                <div class="section">
                    <h2 class="section-title">💡 Öneriler</h2>
                    <div class="alert alert-danger">
                        <span>⚠️</span>
                        <div>
                            <strong>ACİL:</strong> Sızdırmazlık testi prosedürünü revize edin. 
                            Hataların %50'si bu nedenden kaynaklanıyor.
                        </div>
                    </div>
                    <div class="alert alert-warning">
                        <span>📌</span>
                        <div>
                            <strong>YÜKSEK:</strong> Cpk değeri {cap['indices']['Cpk']} seviyesinde. 
                            Süreç iyileştirme projesi başlatılmalı.
                        </div>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        # HTML dosyasını kaydet
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✅ HTML rapor oluşturuldu: {output_path}")
        return output_path
    
    def generate_all_reports(self):
        """
        Tüm rapor tiplerini oluştur
        """
        print("📊 Tüm raporlar oluşturuluyor...")
        
        reports = {}
        
        # 1. Yönetici Özeti
        print("1. Yönetici özeti hazırlanıyor...")
        reports['executive_summary'] = self.generate_executive_summary()
        
        # 2. Haftalık Rapor
        print("2. Haftalık rapor hazırlanıyor...")
        reports['weekly_report'] = self.generate_weekly_report()
        
        # 3. HTML Rapor
        print("3. HTML rapor hazırlanıyor...")
        reports['html_report'] = self.generate_html_report()
        
        print("\n✅ Tüm raporlar başarıyla oluşturuldu!")
        print("📁 Raporlar 'reports/' klasöründe")
        
        return reports
    
    def schedule_reports(self, frequency='weekly'):
        """
        Otomatik rapor zamanlaması (cron job için örnek)
        
        Args:
            frequency: 'daily', 'weekly', 'monthly'
        """
        schedule_config = {
            'daily': '0 8 * * *',  # Her gün saat 08:00
            'weekly': '0 8 * * 1',  # Her Pazartesi 08:00
            'monthly': '0 8 1 * *'  # Her ayın 1'i 08:00
        }
        
        cron_entry = f"{schedule_config.get(frequency)} python {__file__}"
        
        print(f"📅 Rapor zamanlaması için cron entry:")
        print(f"   {cron_entry}")
        print(f"\nBu satırı crontab'a ekleyin: crontab -e")
        
        return cron_entry


# TEST KODU
if __name__ == "__main__":
    print("=== RAPOR OLUŞTURUCU ===")
    
    report_gen = ReportGenerator()
    
    # Tüm raporları oluştur
    all_reports = report_gen.generate_all_reports()
    
    print("\n📊 Oluşturulan Raporlar:")
    for report_type, path in all_reports.items():
        print(f"   - {report_type}: {path}")